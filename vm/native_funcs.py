# Implémentation inline de camelCase (évite la dépendance externe)
class CamelCase:
    def hump(self, s: str) -> str:
        """Convertit 'hello world' en 'helloWorld'."""
        words = s.replace("-", " ").replace("_", " ").split()
        if not words:
            return s
        return words[0].lower() + "".join(w.capitalize() for w in words[1:])
import time as _time
import os
import sys
import random as _random
import math
import platform
import shutil, subprocess, ctypes
try:
    import psutil as _psutil
except ImportError:
    _psutil = None
import json as _json
import sqlite3 as _sqlite3
import difflib as _difflib
import re as _re
import datetime as _datetime
import calendar as _calendar
import csv as _csv
import io as _io
from . ast_nodes import OktopiosMap as _OktopiosMap, OktopiosList as _OktopiosList
try:
    import openpyxl as _openpyxl
except ImportError:
    _openpyxl = None
try:
    import requests as _requests
except ImportError:
    _requests = None
try:
    import pymysql as _pymysql
except ImportError:
    _pymysql = None
try:
    import cv2 as _cv2
except ImportError:
    _cv2 = None
try:
    import numpy as _np
    import librosa as _librosa
except ImportError:
    _np = None
    _librosa = None


# ---------------------------------------------------------------------------
# Reconnaissance faciale (détection + empreinte par histogramme de niveaux
# de gris, comparable via corrélation) — classique/hors-ligne, sans dlib.
# ---------------------------------------------------------------------------

_FACE_CASCADE = None


def _get_face_cascade():
    global _FACE_CASCADE
    if _FACE_CASCADE is None:
        cascade_path = _cv2.data.haarcascades + "haarcascade_frontalface_default.xml"
        _FACE_CASCADE = _cv2.CascadeClassifier(cascade_path)
    return _FACE_CASCADE


def _facial_extract(path):
    if _cv2 is None:
        raise ImportError("opencv n'est pas installé. Lancez : pip install opencv-contrib-python-headless")
    img = _cv2.imread(path)
    if img is None:
        raise ValueError(f"Image illisible : {path}")
    gray = _cv2.cvtColor(img, _cv2.COLOR_BGR2GRAY)
    faces = _get_face_cascade().detectMultiScale(gray, scaleFactor=1.1, minNeighbors=5, minSize=(40, 40))

    results = []
    for (x, y, w, h) in faces:
        face = _cv2.resize(gray[y:y + h, x:x + w], (100, 100))
        hist = _cv2.calcHist([face], [0], None, [256], [0, 256])
        _cv2.normalize(hist, hist)
        results.append(_OktopiosMap({
            "box": _OktopiosList([int(x), int(y), int(w), int(h)]),
            "fingerprint": _OktopiosList([float(v) for v in hist.flatten()]),
        }))
    return _OktopiosList(results)


def _facial_similarity(fp1, fp2):
    if _cv2 is None or _np is None:
        raise ImportError("opencv n'est pas installé. Lancez : pip install opencv-contrib-python-headless")
    a = _np.array(list(fp1), dtype="float32")
    b = _np.array(list(fp2), dtype="float32")
    return float(_cv2.compareHist(a, b, _cv2.HISTCMP_CORREL))


# ---------------------------------------------------------------------------
# Reconnaissance vocale (empreinte MFCC moyenne, comparable par similarité
# cosinus) — classique/hors-ligne, sans moteur de transcription.
# ---------------------------------------------------------------------------

def _vocal_extract(path, n_mfcc=20):
    if _librosa is None:
        raise ImportError("librosa n'est pas installé. Lancez : pip install librosa soundfile")
    y, sr = _librosa.load(path, sr=None, mono=True)
    mfcc = _librosa.feature.mfcc(y=y, sr=sr, n_mfcc=int(n_mfcc))
    fingerprint = _np.mean(mfcc, axis=1)
    return _OktopiosList([float(v) for v in fingerprint])


def _vocal_similarity(fp1, fp2):
    if _np is None:
        raise ImportError("librosa/numpy non installés. Lancez : pip install librosa soundfile")
    a = _np.array(list(fp1), dtype="float64")
    b = _np.array(list(fp2), dtype="float64")
    denom = _np.linalg.norm(a) * _np.linalg.norm(b)
    return float(_np.dot(a, b) / denom) if denom else 0.0


# ---------------------------------------------------------------------------
# Connecteurs de données (json / excel / sql / mysql) — pour __matches_db__
# Les résultats sont enveloppés en OktopiosMap/OktopiosList pour s'intégrer
# nativement au reste du langage (boucles for-in, .get(), affichage, etc.)
# ---------------------------------------------------------------------------

def _wrap_rows(rows):
    return _OktopiosList([_OktopiosMap(dict(r)) for r in rows])


def _read_json(path):
    with open(path, 'r', encoding='utf-8') as f:
        data = _json.load(f)
    if isinstance(data, list):
        return _wrap_rows(data) if data and isinstance(data[0], dict) else _OktopiosList(data)
    if isinstance(data, dict):
        return _OktopiosMap(data)
    return data


def _read_excel(path, sheet=0):
    if _openpyxl is None:
        raise ImportError("openpyxl n'est pas installé. Lancez : pip install openpyxl")
    wb = _openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[sheet] if isinstance(sheet, int) else wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return _OktopiosList([])
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    return _wrap_rows(dict(zip(headers, row)) for row in rows[1:])


def _read_sql(path, query="SELECT name FROM sqlite_master WHERE type='table'"):
    """Lit une base SQLite (.db/.sqlite) ou exécute un script .sql (chargé
    dans une base en mémoire) puis lance `query` et renvoie les lignes."""
    if str(path).lower().endswith(".sql"):
        conn = _sqlite3.connect(":memory:")
        with open(path, "r", encoding="utf-8") as f:
            conn.executescript(f.read())
    else:
        conn = _sqlite3.connect(path)
    conn.row_factory = _sqlite3.Row
    rows = [dict(row) for row in conn.execute(query).fetchall()]
    conn.close()
    return _wrap_rows(rows)


def _read_mysql(host, user, password, database, query, port=3306):
    if _pymysql is None:
        raise ImportError("pymysql n'est pas installé. Lancez : pip install pymysql")
    conn = _pymysql.connect(
        host=host, user=user, password=password, database=database,
        port=int(port), cursorclass=_pymysql.cursors.DictCursor,
    )
    try:
        with conn.cursor() as cur:
            cur.execute(query)
            return _wrap_rows(cur.fetchall())
    finally:
        conn.close()


def _unwrap(value):
    """Convertit OktopiosMap/OktopiosList (récursivement) en dict/list Python
    natifs, pour pouvoir les sérialiser (json/excel/sql/mysql)."""
    if isinstance(value, _OktopiosMap):
        entries = value.entries if value.entries is not None else dict(value)
        return {k: _unwrap(v) for k, v in entries.items()}
    if isinstance(value, _OktopiosList):
        elements = value.elements if value.elements is not None else list(value)
        return [_unwrap(v) for v in elements]
    if isinstance(value, dict):
        return {k: _unwrap(v) for k, v in value.items()}
    if isinstance(value, (list, tuple)):
        return [_unwrap(v) for v in value]
    return value


def _as_rows(data):
    rows = _unwrap(data)
    if isinstance(rows, dict):
        rows = [rows]
    return rows or []


def _write_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(_unwrap(data), f, ensure_ascii=False, indent=2)
    return True


def _write_excel(path, data, sheet_name="Sheet1"):
    if _openpyxl is None:
        raise ImportError("openpyxl n'est pas installé. Lancez : pip install openpyxl")
    rows = _as_rows(data)
    wb = _openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    if rows:
        headers = list(rows[0].keys())
        ws.append(headers)
        for row in rows:
            ws.append([row.get(h) for h in headers])
    wb.save(path)
    return True


def _write_sql(path, table, data):
    rows = _as_rows(data)
    if not rows:
        return False
    columns = list(rows[0].keys())
    conn = _sqlite3.connect(path)
    col_defs = ", ".join(f'"{c}"' for c in columns)
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({col_defs})')
    placeholders = ", ".join("?" for _ in columns)
    for row in rows:
        conn.execute(
            f'INSERT INTO "{table}" ({col_defs}) VALUES ({placeholders})',
            [row.get(c) for c in columns],
        )
    conn.commit()
    conn.close()
    return True


def _sql_type_for(value):
    if isinstance(value, bool):
        return "BOOLEAN"
    if isinstance(value, int):
        return "BIGINT"
    if isinstance(value, float):
        return "DOUBLE"
    return "TEXT"


def _write_mysql(host, user, password, database, table, data, port=3306):
    if _pymysql is None:
        raise ImportError("pymysql n'est pas installé. Lancez : pip install pymysql")
    rows = _as_rows(data)
    if not rows:
        return False
    columns = list(rows[0].keys())
    conn = _pymysql.connect(host=host, user=user, password=password, database=database, port=int(port))
    try:
        with conn.cursor() as cur:
            # Crée la table si elle n'existe pas encore (contrairement à SQLite,
            # MySQL ne le fait jamais implicitement — bug trouvé en testant
            # contre un vrai serveur MariaDB).
            col_defs_create = ", ".join(
                f"`{c}` {_sql_type_for(rows[0][c])}" for c in columns
            )
            cur.execute(f"CREATE TABLE IF NOT EXISTS `{table}` ({col_defs_create})")

            col_list = ", ".join(f"`{c}`" for c in columns)
            placeholders = ", ".join(["%s"] * len(columns))
            cur.executemany(
                f"INSERT INTO `{table}` ({col_list}) VALUES ({placeholders})",
                [tuple(row.get(c) for c in columns) for row in rows],
            )
        conn.commit()
        return True
    finally:
        conn.close()


def _text_normalize(s):
    return str(s).strip().lower()


def _text_similarity(a, b):
    return _difflib.SequenceMatcher(None, _text_normalize(a), _text_normalize(b)).ratio()


# ---------------------------------------------------------------------------
# IA Adaptive Module (Heart 3) — Ollama (local) / DeepSeek / StarCoder
# ---------------------------------------------------------------------------

def _retry_call(fn, retries=2, backoff=1.0):
    """Réessaie fn() en cas d'échec TRANSITOIRE (connexion, timeout, 5xx),
    jamais sur une erreur permanente (401/403/404) — inutile de réessayer
    une clé API invalide. Backoff exponentiel entre les tentatives."""
    import time as _t
    last_exc = None
    for attempt in range(retries + 1):
        try:
            return fn()
        except Exception as e:
            last_exc = e
            msg = str(e)
            permanent = any(code in msg for code in (" 401", " 403", " 404", "Unauthorized", "Forbidden"))
            if permanent or attempt == retries:
                raise
            _t.sleep(backoff * (2 ** attempt))
    raise last_exc


def _require_requests():
    if _requests is None:
        raise ImportError(
            "Le module Python 'requests' n'est pas installé. "
            "Lancez : pip install requests (nécessaire pour IAModule.*)"
        )
    return _requests


def _ia_ollama(prompt, model="llama3", host="http://localhost:11434", timeout=60, retries=2):
    """Appelle un serveur Ollama LOCAL (à lancer côté Ali : `ollama serve`).
    CodeLlama et Mistral n'ont pas besoin d'une fonction séparée : ce sont
    juste des modèles Ollama, ex. model='codellama' ou model='mistral'
    (après `ollama pull codellama` / `ollama pull mistral`).
    Aucune restriction réseau ici puisque c'est du localhost, mais ça
    nécessite que le serveur tourne réellement sur la machine qui exécute
    le programme Oktopios."""
    requests = _require_requests()

    def _do():
        resp = requests.post(
            f"{host}/api/generate",
            json={"model": model, "prompt": prompt, "stream": False},
            timeout=timeout,
        )
        resp.raise_for_status()
        return resp.json().get("response", "")

    try:
        return _retry_call(_do, retries=retries)
    except Exception as e:
        raise Exception(f"[IAModule.ollama] Impossible de contacter Ollama sur {host} (modèle '{model}') après {retries + 1} tentative(s) : {e}")


def _ia_ollama_stream(prompt, model, host, timeout, on_chunk):
    """Variante streaming : appelle on_chunk(texte_partiel) à chaque morceau
    reçu d'Ollama, et renvoie le texte complet accumulé à la fin. Garde la
    connexion "vivante" sur les génrations longues plutôt qu'une seule
    attente bloquante."""
    requests = _require_requests()
    import json as _json
    full_text = []
    with requests.post(
        f"{host}/api/generate",
        json={"model": model, "prompt": prompt, "stream": True},
        timeout=timeout,
        stream=True,
    ) as resp:
        resp.raise_for_status()
        for line in resp.iter_lines():
            if not line:
                continue
            chunk = _json.loads(line)
            piece = chunk.get("response", "")
            if piece:
                full_text.append(piece)
                if on_chunk is not None:
                    on_chunk(piece)
            if chunk.get("done"):
                break
    return "".join(full_text)


def _ia_deepseek(prompt, api_key=None, model=None, timeout=60, retries=2, host="http://localhost:11434"):
    """DeepSeek est open source (licence MIT pour R1) et disponible
    directement via Ollama — donc PAR DÉFAUT (pas de api_key fournie), on
    route vers Ollama local, gratuit, sans clé : `ollama pull deepseek-r1`.
    Si une api_key EST fournie, on utilise l'API hébergée payante à la place
    (utile si la machine ne peut pas faire tourner le modèle localement)."""
    if not api_key:
        return _ia_ollama(prompt, model or "deepseek-r1", host, timeout, retries)

    requests = _require_requests()

    def _do():
        resp = requests.post(
            "https://api.deepseek.com/chat/completions",
            headers={"Authorization": f"Bearer {api_key}"},
            json={"model": model or "deepseek-chat", "messages": [{"role": "user", "content": prompt}]},
            timeout=timeout,
        )
        resp.raise_for_status()
        return resp.json()["choices"][0]["message"]["content"]

    try:
        return _retry_call(_do, retries=retries)
    except Exception as e:
        raise Exception(f"[IAModule.deepseek] Erreur API DeepSeek après {retries + 1} tentative(s) : {e}")


def _ia_starcoder(prompt, api_key=None, model=None, timeout=60, retries=2, host="http://localhost:11434"):
    """StarCoder2 est open source (licence BigCode OpenRAIL-M) et disponible
    directement via Ollama — donc PAR DÉFAUT (pas de api_key fournie), on
    route vers Ollama local, gratuit, sans clé : `ollama pull starcoder2`.
    Si une api_key EST fournie, on utilise l'API d'inférence Hugging Face
    à la place (utile si la machine ne peut pas faire tourner le modèle
    localement — vérifier que le modèle visé est encore dispo en serverless)."""
    if not api_key:
        return _ia_ollama(prompt, model or "starcoder2", host, timeout, retries)

    requests = _require_requests()

    def _do():
        resp = requests.post(
            f"https://api-inference.huggingface.co/models/{model or 'bigcode/starcoder2-15b'}",
            headers={"Authorization": f"Bearer {api_key}"},
            json={"inputs": prompt},
            timeout=timeout,
        )
        resp.raise_for_status()
        data = resp.json()
        if isinstance(data, list) and data:
            return data[0].get("generated_text", "")
        return str(data)

    try:
        return _retry_call(_do, retries=retries)
    except Exception as e:
        raise Exception(f"[IAModule.starcoder] Erreur API StarCoder après {retries + 1} tentative(s) : {e}")




# ---------------------------------------------------------------------------
# Helpers fichiers — fermeture garantie via with (évite les fuites de handles)
# ---------------------------------------------------------------------------

def _file_read(path: str) -> str:
    with open(path, 'r', encoding='utf-8') as f:
        return f.read()

def _file_write(path: str, content) -> None:
    with open(path, 'w', encoding='utf-8') as f:
        f.write(str(content))

def _file_append(path: str, content) -> None:
    with open(path, 'a', encoding='utf-8') as f:
        f.write(str(content))

def _file_readlines(path: str):
    with open(path, 'r', encoding='utf-8') as f:
        return f.read().splitlines()

def _file_hash(path: str, algo: str = "sha256") -> str:
    import hashlib
    h = hashlib.new(algo)
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Helpers pour le namespace Hash — hachage cryptographique & encodage Base64
# ---------------------------------------------------------------------------
import hashlib as _hashlib
import hmac as _hmac
import base64 as _base64

def _hash_digest(algo: str, s: str) -> str:
    """Retourne le digest hexadécimal de la chaîne `s` avec l'algorithme `algo`."""
    return _hashlib.new(algo, str(s).encode("utf-8")).hexdigest()

def _hash_hmac(key: str, msg: str, algo: str = "sha256") -> str:
    """Retourne le HMAC-<algo> (hex) du message `msg` signé avec `key`."""
    return _hmac.new(
        str(key).encode("utf-8"),
        str(msg).encode("utf-8"),
        digestmod=_hashlib.new(algo).name
    ).hexdigest()

def _hash_b64encode(s: str) -> str:
    return _base64.b64encode(str(s).encode("utf-8")).decode("ascii")

def _hash_b64decode(s: str) -> str:
    return _base64.b64decode(str(s).encode("ascii")).decode("utf-8")

def _hash_b64url_encode(s: str) -> str:
    return _base64.urlsafe_b64encode(str(s).encode("utf-8")).decode("ascii")

def _hash_b64url_decode(s: str) -> str:
    # Ajouter le padding manquant si nécessaire
    data = str(s).encode("ascii")
    data += b"=" * (-len(data) % 4)
    return _base64.urlsafe_b64decode(data).decode("utf-8")

def _hash_compare(h1: str, h2: str) -> bool:
    """Comparaison en temps constant (résistant aux attaques temporelles)."""
    return _hmac.compare_digest(str(h1), str(h2))


# ---------------------------------------------------------------------------
# Helpers pour le namespace List et Type
# ---------------------------------------------------------------------------

def _list_flatten(lst):
    """Aplatit récursivement une liste imbriquée."""
    result = []
    for item in lst:
        if isinstance(item, (list, _OktopiosList)):
            result.extend(_list_flatten(list(item)))
        else:
            result.append(item)
    return result


def _list_product(lst):
    """Produit de tous les éléments d'une liste."""
    result = 1
    for x in lst:
        result *= x
    return result


def _list_rotate(lst, n):
    """Rotation de la liste de n positions vers la gauche (n<0 = droite)."""
    if not lst:
        return lst
    n = n % len(lst)
    return lst[n:] + lst[:n]


def _oktype(x) -> str:
    """Retourne le nom de type Oktopios d'une valeur."""
    if x is None:
        return "null"
    if isinstance(x, bool):
        return "bool"
    if isinstance(x, int):
        return "int"
    if isinstance(x, float):
        return "float"
    if isinstance(x, str):
        return "string"
    if isinstance(x, _OktopiosList):
        return "list"
    if isinstance(x, _OktopiosMap):
        return "map"
    if isinstance(x, list):
        return "list"
    if isinstance(x, dict):
        return "map"
    # Retourne le nom de classe pour les types personnalisés (Tentacle, etc.)
    return type(x).__name__



# ---------------------------------------------------------------------------
# Regex helpers
# ---------------------------------------------------------------------------

def _re_groups(pattern, s):
    """Return capture groups from the first match, or empty list."""
    m = _re.search(pattern, s)
    if m is None:
        return _OktopiosList([])
    return _OktopiosList(list(m.groups()))

def _re_find_all(pattern, s):
    """Return all non-overlapping matches as a list of strings."""
    return _OktopiosList(_re.findall(pattern, s))

def _re_named_groups(pattern, s):
    """Return a map of named capture groups from the first match."""
    m = _re.search(pattern, s)
    if m is None:
        return _OktopiosMap({})
    return _OktopiosMap(m.groupdict())


# ---------------------------------------------------------------------------
# Date — arithmétique et manipulation de dates (namespace Date)
# Les dates sont manipulées sous forme de chaînes ISO "YYYY-MM-DD".
# Toutes les fonctions acceptent aussi des chaînes comme "DD/MM/YYYY".
# ---------------------------------------------------------------------------

_DATE_FORMATS = [
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y",
    "%d-%m-%Y", "%Y%m%d", "%Y/%m/%d",
    "%d %B %Y", "%d %b %Y",
]


def _date_to_dt(d) -> _datetime.datetime:
    """Convertit une chaîne de date en objet datetime (pour calculs internes)."""
    if isinstance(d, _datetime.datetime):
        return d
    if isinstance(d, _datetime.date):
        return _datetime.datetime(d.year, d.month, d.day)
    s = str(d).strip()
    for fmt in _DATE_FORMATS:
        try:
            return _datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    raise ValueError(
        f"Date non reconnue : {d!r}. "
        "Formats acceptés : YYYY-MM-DD, DD/MM/YYYY, MM/DD/YYYY, ..."
    )


def _date_parse(s, fmt="%Y-%m-%d") -> str:
    """Parse une date avec un format donné, retourne ISO YYYY-MM-DD."""
    return _datetime.datetime.strptime(str(s).strip(), str(fmt)).strftime("%Y-%m-%d")


def _date_format(d, fmt="%d/%m/%Y") -> str:
    """Formate une date ISO vers un format arbitraire."""
    return _date_to_dt(d).strftime(str(fmt))


def _date_add(d, n, unit="days") -> str:
    """Ajoute n unités (days/weeks/months/years) à une date, retourne ISO."""
    dt = _date_to_dt(d)
    n = int(n)
    unit = str(unit).lower().rstrip("s")  # normalise "days" → "day"
    if unit == "day":
        result = dt + _datetime.timedelta(days=n)
    elif unit == "week":
        result = dt + _datetime.timedelta(weeks=n)
    elif unit == "month":
        month = dt.month + n
        year = dt.year + (month - 1) // 12
        month = (month - 1) % 12 + 1
        day = min(dt.day, _calendar.monthrange(year, month)[1])
        result = dt.replace(year=year, month=month, day=day)
    elif unit == "year":
        try:
            result = dt.replace(year=dt.year + n)
        except ValueError:  # 29 fév en année non-bissextile
            result = dt.replace(year=dt.year + n, day=28)
    else:
        raise ValueError(
            f"Unité inconnue : {unit!r}. Valeurs valides : 'days', 'weeks', 'months', 'years'."
        )
    return result.strftime("%Y-%m-%d")


def _date_diff(d1, d2, unit="days") -> int:
    """Retourne la différence entre deux dates dans l'unité demandée."""
    dt1 = _date_to_dt(d1)
    dt2 = _date_to_dt(d2)
    delta = dt2 - dt1
    unit = str(unit).lower().rstrip("s")
    if unit == "day":
        return delta.days
    if unit == "week":
        return delta.days // 7
    if unit == "hour":
        return int(delta.total_seconds() // 3600)
    if unit == "minute":
        return int(delta.total_seconds() // 60)
    if unit == "second":
        return int(delta.total_seconds())
    return delta.days  # défaut : jours


def _date_compare(d1, d2) -> int:
    """Retourne -1, 0 ou 1 selon que d1 est avant, égal ou après d2."""
    dt1 = _date_to_dt(d1)
    dt2 = _date_to_dt(d2)
    return 0 if dt1 == dt2 else (-1 if dt1 < dt2 else 1)


def _date_is_leap_year(year: int) -> bool:
    """Retourne true si l'année est bissextile."""
    y = int(year)
    return (y % 4 == 0 and y % 100 != 0) or (y % 400 == 0)


def _date_days_in_month(year: int, month: int) -> int:
    """Retourne le nombre de jours dans un mois donné."""
    return _calendar.monthrange(int(year), int(month))[1]


_WEEKDAY_EN = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
_WEEKDAY_FR = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]


def _date_weekday(d) -> int:
    """Retourne le numéro du jour de la semaine (0=Lundi … 6=Dimanche)."""
    return _date_to_dt(d).weekday()


def _date_weekday_name(d, lang="en") -> str:
    """Retourne le nom du jour en anglais (par défaut) ou en français ('fr')."""
    wd = _date_to_dt(d).weekday()
    if str(lang).lower() in ("fr", "french", "français"):
        return _WEEKDAY_FR[wd]
    return _WEEKDAY_EN[wd]


def _date_to_timestamp(d) -> int:
    """Convertit une date en timestamp Unix (entier)."""
    return int(_date_to_dt(d).timestamp())


def _date_from_timestamp(ts) -> str:
    """Convertit un timestamp Unix en date ISO YYYY-MM-DD."""
    return _datetime.datetime.fromtimestamp(float(ts)).strftime("%Y-%m-%d")


# ---------------------------------------------------------------------------
# Helpers Json
# ---------------------------------------------------------------------------

def _okp_to_python(v):
    """Convertit récursivement une valeur Oktopios en type Python natif."""
    if isinstance(v, _OktopiosMap):
        return {str(k): _okp_to_python(val) for k, val in v.items()}
    if isinstance(v, (_OktopiosList, list)):
        return [_okp_to_python(x) for x in v]
    return v


def _python_to_okp(v):
    """Convertit récursivement un type Python (dict/list) en valeur Oktopios."""
    if isinstance(v, dict):
        return _OktopiosMap({k: _python_to_okp(val) for k, val in v.items()})
    if isinstance(v, list):
        return _OktopiosList([_python_to_okp(x) for x in v])
    return v


def _json_parse(s: str):
    """Désérialise une chaîne JSON en valeur Oktopios."""
    return _python_to_okp(_json.loads(s))


def _json_stringify(v, indent=None) -> str:
    """Sérialise une valeur Oktopios en chaîne JSON."""
    return _json.dumps(_okp_to_python(v), ensure_ascii=False,
                       indent=int(indent) if indent is not None else None)


def _json_is_valid(s: str) -> bool:
    """Retourne True si la chaîne est du JSON valide."""
    try:
        _json.loads(s)
        return True
    except _json.JSONDecodeError:
        return False


def _json_path_get(obj, path: str, default=None):
    """Lit une valeur à un chemin 'a.b.c' dans un objet imbriqué."""
    keys = path.split(".")
    cur = obj
    for k in keys:
        if isinstance(cur, (dict, _OktopiosMap)):
            cur = dict(cur).get(k, None)
            if cur is None:
                return default
        elif isinstance(cur, (_OktopiosList, list)):
            try:
                cur = list(cur)[int(k)]
            except (ValueError, IndexError):
                return default
        else:
            return default
    return cur


def _json_path_has(obj, path: str) -> bool:
    """Retourne True si le chemin 'a.b.c' existe dans l'objet."""
    sentinel = object()
    return _json_path_get(obj, path, sentinel) is not sentinel


def _json_path_set(obj, path: str, val):
    """Retourne un nouvel objet avec la valeur modifiée au chemin 'a.b.c'."""
    keys = path.split(".")
    if not keys:
        return val

    def _set_recursive(cur, ks, v):
        k = ks[0]
        rest = ks[1:]
        if isinstance(cur, (dict, _OktopiosMap)):
            d = dict(cur)
            d[k] = _set_recursive(d.get(k, _OktopiosMap({})), rest, v) if rest else v
            return _OktopiosMap(d)
        return val

    return _set_recursive(obj, keys, val)


def _json_deep_merge(a, b):
    """Merge profond de deux objets Oktopios (b écrase a en cas de conflit)."""
    if isinstance(a, (dict, _OktopiosMap)) and isinstance(b, (dict, _OktopiosMap)):
        result = dict(a).copy()
        for k, v in dict(b).items():
            result[k] = _json_deep_merge(result[k], v) if k in result else v
        return _OktopiosMap(result)
    return b


def _json_from_file(path: str):
    """Charge un fichier JSON et retourne une valeur Oktopios."""
    with open(path, "r", encoding="utf-8") as f:
        return _python_to_okp(_json.load(f))


def _json_to_file(path: str, v, indent: int = 2) -> None:
    """Écrit une valeur Oktopios dans un fichier JSON."""
    with open(path, "w", encoding="utf-8") as f:
        _json.dump(_okp_to_python(v), f, ensure_ascii=False, indent=indent)


# ---------------------------------------------------------------------------
# Http — requêtes HTTP (GET / POST / PUT / DELETE / PATCH)
# Nécessite le module `requests` (pip install requests ou oktopios[ia]).
# Chaque méthode retourne un OktopiosMap :
#   { "status": int, "ok": bool, "body": str, "headers": OktopiosMap }
# ---------------------------------------------------------------------------

def _http_require_requests():
    if _requests is None:
        raise RuntimeError(
            "Le namespace Http nécessite le module 'requests'.\n"
            "Installez-le avec : pip install requests\n"
            "ou : pip install oktopios[ia]"
        )

def _http_response_to_okp(resp) -> _OktopiosMap:
    """Convertit un objet requests.Response en OktopiosMap Oktopios."""
    return _OktopiosMap({
        "status":  resp.status_code,
        "ok":      resp.ok,
        "body":    resp.text,
        "headers": _OktopiosMap(dict(resp.headers)),
    })

def _http_build_headers(headers) -> dict:
    """Accepte un OktopiosMap ou un dict Python et retourne un dict Python."""
    if headers is None:
        return {}
    if isinstance(headers, _OktopiosMap):
        return dict(headers)
    if isinstance(headers, dict):
        return headers
    return {}

def _http_get(url: str, headers=None, timeout: int = 30) -> _OktopiosMap:
    _http_require_requests()
    resp = _requests.get(str(url), headers=_http_build_headers(headers), timeout=int(timeout))
    return _http_response_to_okp(resp)

def _http_post(url: str, body=None, headers=None, json_body=None, timeout: int = 30) -> _OktopiosMap:
    _http_require_requests()
    h = _http_build_headers(headers)
    if json_body is not None:
        resp = _requests.post(str(url), json=_okp_to_python(json_body), headers=h, timeout=int(timeout))
    elif body is not None:
        resp = _requests.post(str(url), data=str(body), headers=h, timeout=int(timeout))
    else:
        resp = _requests.post(str(url), headers=h, timeout=int(timeout))
    return _http_response_to_okp(resp)

def _http_put(url: str, body=None, headers=None, json_body=None, timeout: int = 30) -> _OktopiosMap:
    _http_require_requests()
    h = _http_build_headers(headers)
    if json_body is not None:
        resp = _requests.put(str(url), json=_okp_to_python(json_body), headers=h, timeout=int(timeout))
    elif body is not None:
        resp = _requests.put(str(url), data=str(body), headers=h, timeout=int(timeout))
    else:
        resp = _requests.put(str(url), headers=h, timeout=int(timeout))
    return _http_response_to_okp(resp)

def _http_patch(url: str, body=None, headers=None, json_body=None, timeout: int = 30) -> _OktopiosMap:
    _http_require_requests()
    h = _http_build_headers(headers)
    if json_body is not None:
        resp = _requests.patch(str(url), json=_okp_to_python(json_body), headers=h, timeout=int(timeout))
    elif body is not None:
        resp = _requests.patch(str(url), data=str(body), headers=h, timeout=int(timeout))
    else:
        resp = _requests.patch(str(url), headers=h, timeout=int(timeout))
    return _http_response_to_okp(resp)

def _http_delete(url: str, headers=None, timeout: int = 30) -> _OktopiosMap:
    _http_require_requests()
    resp = _requests.delete(str(url), headers=_http_build_headers(headers), timeout=int(timeout))
    return _http_response_to_okp(resp)

def _http_json(response) -> object:
    """Parse le corps de la réponse comme JSON et retourne une valeur Oktopios."""
    body = dict(response).get("body", "")
    return _python_to_okp(_json.loads(str(body)))

def _http_status(response) -> int:
    return int(dict(response).get("status", 0))

def _http_ok(response) -> bool:
    return bool(dict(response).get("ok", False))

def _http_body(response) -> str:
    return str(dict(response).get("body", ""))

def _http_headers(response) -> _OktopiosMap:
    h = dict(response).get("headers", _OktopiosMap({}))
    return h if isinstance(h, _OktopiosMap) else _OktopiosMap(dict(h))


# ---------------------------------------------------------------------------
# Queue — file d'attente FIFO (premier entré, premier sorti)
# Représentation interne : OktopiosMap { "_t": "Q", "_d": [items...] }
# Les mutations (enqueue/dequeue/clear) modifient la liste interne en place.
# ---------------------------------------------------------------------------

def _queue_create(lst=None):
    items = list(lst) if lst is not None else []
    return _OktopiosMap({"_t": "Q", "_d": items})

def _queue_enqueue(q, val):
    dict.__getitem__(q, "_d").append(val)
    return None

def _queue_dequeue(q):
    items = dict.__getitem__(q, "_d")
    if not items:
        raise Exception("[Queue] Impossible de déqueuer : la file est vide")
    return items.pop(0)

def _queue_peek(q):
    items = dict.__getitem__(q, "_d")
    if not items:
        raise Exception("[Queue] Impossible de lire : la file est vide")
    return items[0]

def _queue_size(q):
    return len(dict.__getitem__(q, "_d"))

def _queue_isEmpty(q):
    return len(dict.__getitem__(q, "_d")) == 0

def _queue_toList(q):
    return _OktopiosList(list(dict.__getitem__(q, "_d")))

def _queue_clear(q):
    dict.__getitem__(q, "_d").clear()
    return None

def _queue_contains(q, val):
    return val in dict.__getitem__(q, "_d")


# ---------------------------------------------------------------------------
# Stack — pile LIFO (dernier entré, premier sorti)
# Représentation interne : OktopiosMap { "_t": "S", "_d": [items...] }
# Le sommet (top) est le dernier élément de la liste.
# Les mutations (push/pop/clear) modifient la liste interne en place.
# ---------------------------------------------------------------------------

def _stack_create(lst=None):
    items = list(lst) if lst is not None else []
    return _OktopiosMap({"_t": "S", "_d": items})

def _stack_push(s, val):
    dict.__getitem__(s, "_d").append(val)
    return None

def _stack_pop(s):
    items = dict.__getitem__(s, "_d")
    if not items:
        raise Exception("[Stack] Impossible de dépiler : la pile est vide")
    return items.pop()

def _stack_peek(s):
    items = dict.__getitem__(s, "_d")
    if not items:
        raise Exception("[Stack] Impossible de lire : la pile est vide")
    return items[-1]

def _stack_size(s):
    return len(dict.__getitem__(s, "_d"))

def _stack_isEmpty(s):
    return len(dict.__getitem__(s, "_d")) == 0

def _stack_toList(s):
    return _OktopiosList(list(reversed(dict.__getitem__(s, "_d"))))

def _stack_clear(s):
    dict.__getitem__(s, "_d").clear()
    return None

def _stack_contains(s, val):
    return val in dict.__getitem__(s, "_d")


# ---------------------------------------------------------------------------
# Helpers — Namespace Stats
# Bibliothèque standard uniquement : statistics, math
# ---------------------------------------------------------------------------
import statistics as _statistics

def _stats_to_floats(lst):
    """Convert OktopiosList / Python list to a list of floats, raise on empty."""
    data = list(lst)
    if not data:
        raise ValueError("Stats: liste vide")
    return [float(x) for x in data]

def _stats_mean(lst):
    return _statistics.mean(_stats_to_floats(lst))

def _stats_median(lst):
    return _statistics.median(_stats_to_floats(lst))

def _stats_mode(lst):
    try:
        return _statistics.mode(_stats_to_floats(lst))
    except _statistics.StatisticsError:
        # Python <3.8 raises on multimodal; return first element of most common
        data = _stats_to_floats(lst)
        from collections import Counter
        return Counter(data).most_common(1)[0][0]

def _stats_geomean(lst):
    data = _stats_to_floats(lst)
    if any(x <= 0 for x in data):
        raise ValueError("Stats.geomean: toutes les valeurs doivent être > 0")
    return math.exp(sum(math.log(x) for x in data) / len(data))

def _stats_harmean(lst):
    data = _stats_to_floats(lst)
    if any(x <= 0 for x in data):
        raise ValueError("Stats.harmean: toutes les valeurs doivent être > 0")
    return len(data) / sum(1.0 / x for x in data)

def _stats_variance(lst, pop=False):
    data = _stats_to_floats(lst)
    if len(data) < 2 and not pop:
        raise ValueError("Stats.variance: au moins 2 valeurs requises (variance échantillon)")
    return _statistics.pvariance(data) if pop else _statistics.variance(data)

def _stats_stddev(lst, pop=False):
    data = _stats_to_floats(lst)
    if len(data) < 2 and not pop:
        raise ValueError("Stats.stddev: au moins 2 valeurs requises (écart-type échantillon)")
    return _statistics.pstdev(data) if pop else _statistics.stdev(data)

def _stats_range(lst):
    data = _stats_to_floats(lst)
    return max(data) - min(data)

def _stats_sorted_floats(lst):
    return sorted(_stats_to_floats(lst))

def _stats_percentile(lst, p):
    """p in [0, 100] — interpolation linéaire (méthode exclusive)."""
    data = _stats_sorted_floats(lst)
    p = float(p)
    if not (0 <= p <= 100):
        raise ValueError("Stats.percentile: p doit être entre 0 et 100")
    n = len(data)
    if n == 1:
        return data[0]
    idx = (p / 100) * (n - 1)
    lo, hi = int(math.floor(idx)), int(math.ceil(idx))
    if lo == hi:
        return data[lo]
    return data[lo] + (data[hi] - data[lo]) * (idx - lo)

def _stats_quartiles(lst):
    q1 = _stats_percentile(lst, 25)
    q2 = _stats_percentile(lst, 50)
    q3 = _stats_percentile(lst, 75)
    return _OktopiosList([q1, q2, q3])

def _stats_iqr(lst):
    return _stats_percentile(lst, 75) - _stats_percentile(lst, 25)

def _stats_mad(lst):
    """Écart absolu médian."""
    data = _stats_to_floats(lst)
    med = _statistics.median(data)
    return _statistics.median([abs(x - med) for x in data])

def _stats_normalize(lst):
    """Normalisation min-max vers [0, 1]."""
    data = _stats_to_floats(lst)
    lo, hi = min(data), max(data)
    if hi == lo:
        return _OktopiosList([0.0] * len(data))
    return _OktopiosList([(x - lo) / (hi - lo) for x in data])

def _stats_zscore(lst):
    """Z-scores (écarts centrés réduits)."""
    data = _stats_to_floats(lst)
    if len(data) < 2:
        raise ValueError("Stats.zscore: au moins 2 valeurs requises")
    mu = _statistics.mean(data)
    sigma = _statistics.stdev(data)
    if sigma == 0:
        return _OktopiosList([0.0] * len(data))
    return _OktopiosList([(x - mu) / sigma for x in data])

def _stats_covariance(a, b):
    da = _stats_to_floats(a)
    db = _stats_to_floats(b)
    if len(da) != len(db):
        raise ValueError("Stats.covariance: les deux listes doivent avoir la même taille")
    n = len(da)
    if n < 2:
        raise ValueError("Stats.covariance: au moins 2 paires requises")
    mu_a, mu_b = _statistics.mean(da), _statistics.mean(db)
    return sum((x - mu_a) * (y - mu_b) for x, y in zip(da, db)) / (n - 1)

def _stats_correlation(a, b):
    cov = _stats_covariance(a, b)
    sa = _statistics.stdev(_stats_to_floats(a))
    sb = _statistics.stdev(_stats_to_floats(b))
    if sa == 0 or sb == 0:
        return 0.0
    return cov / (sa * sb)

def _stats_describe(lst):
    """Retourne un résumé statistique sous forme d'OktopiosMap."""
    data = _stats_to_floats(lst)
    n = len(data)
    q1 = _stats_percentile(lst, 25)
    q3 = _stats_percentile(lst, 75)
    entries = {
        "count":    float(n),
        "sum":      float(sum(data)),
        "min":      float(min(data)),
        "max":      float(max(data)),
        "range":    float(max(data) - min(data)),
        "mean":     _statistics.mean(data),
        "median":   _statistics.median(data),
        "variance": _statistics.variance(data) if n >= 2 else 0.0,
        "stddev":   _statistics.stdev(data) if n >= 2 else 0.0,
        "q1":       q1,
        "q3":       q3,
        "iqr":      q3 - q1,
    }
    return _OktopiosMap(entries)


# ---------------------------------------------------------------------------
# Namespace Csv — lecture, écriture et conversion CSV (stdlib uniquement)
# Utilise uniquement le module `csv` de la bibliothèque standard Python.
# ---------------------------------------------------------------------------

def _csv_to_list(rows, has_header):
    """Convertit des lignes brutes en OktopiosList de OktopiosMap (avec en-tête)
    ou OktopiosList de OktopiosList (sans en-tête)."""
    if not rows:
        return _OktopiosList([])
    if has_header:
        header = rows[0]
        result = []
        for row in rows[1:]:
            entry = _OktopiosMap({header[i]: (row[i] if i < len(row) else "") for i in range(len(header))})
            result.append(entry)
        return _OktopiosList(result)
    else:
        return _OktopiosList([_OktopiosList(row) for row in rows])


def _csv_read(path, delimiter=",", has_header=True):
    """Lit un fichier CSV et retourne une liste de maps (avec en-tête)
    ou une liste de listes (sans en-tête)."""
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        reader = _csv.reader(f, delimiter=str(delimiter))
        rows = list(reader)
    return _csv_to_list(rows, bool(has_header))


def _csv_write(path, data, delimiter=",", header=None):
    """Écrit des données dans un fichier CSV.
    data peut être une liste de maps ou une liste de listes."""
    rows = []
    # Détecter si c'est une liste de maps (OktopiosMap ou dict)
    if data and isinstance(data[0] if not hasattr(data, 'values') else list(data.values)[0], (_OktopiosMap, dict)):
        items = data.values if hasattr(data, 'values') else data
        first = items[0] if not hasattr(items, '__getitem__') else items[0]
        keys = list(first.values.keys()) if hasattr(first, 'values') else list(first.keys())
        if header is None:
            header = keys
        rows.append(header)
        for item in (items.values if hasattr(items, 'values') else items):
            d = item.values if hasattr(item, 'values') else item
            rows.append([str(d.get(k, "")) for k in header])
    else:
        items = data.values if hasattr(data, 'values') else data
        if header:
            rows.append(header)
        for row in items:
            raw = row.values if hasattr(row, 'values') else row
            rows.append([str(v) for v in (raw if not isinstance(raw, str) else [raw])])
    with open(str(path), "w", newline="", encoding="utf-8") as f:
        w = _csv.writer(f, delimiter=str(delimiter))
        w.writerows(rows)
    return True


def _csv_parse(text, delimiter=",", has_header=True):
    """Analyse un texte CSV et retourne une liste de maps ou de listes."""
    reader = _csv.reader(_io.StringIO(str(text)), delimiter=str(delimiter))
    rows = list(reader)
    return _csv_to_list(rows, bool(has_header))


def _csv_stringify(data, delimiter=",", header=None):
    """Convertit des données en texte CSV et le retourne sous forme de chaîne."""
    buf = _io.StringIO()
    rows = []
    items = data.values if hasattr(data, 'values') else data
    if items and isinstance(items[0] if not hasattr(items, '__getitem__') else items[0], (_OktopiosMap, dict)):
        first = items[0] if not hasattr(items, '__getitem__') else items[0]
        keys = list(first.values.keys()) if hasattr(first, 'values') else list(first.keys())
        if header is None:
            header = keys
        rows.append(header)
        for item in items:
            d = item.values if hasattr(item, 'values') else item
            rows.append([str(d.get(k, "")) for k in header])
    else:
        if header:
            rows.append(header)
        for row in items:
            raw = row.values if hasattr(row, 'values') else row
            rows.append([str(v) for v in (raw if not isinstance(raw, str) else [raw])])
    w = _csv.writer(buf, delimiter=str(delimiter))
    w.writerows(rows)
    return buf.getvalue()


def _csv_head(path, n=5, delimiter=","):
    """Retourne les n premières lignes du fichier CSV sous forme de liste de listes."""
    rows = []
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        reader = _csv.reader(f, delimiter=str(delimiter))
        for i, row in enumerate(reader):
            if i >= int(n):
                break
            rows.append(_OktopiosList(row))
    return _OktopiosList(rows)


def _csv_columns(path, delimiter=","):
    """Retourne la première ligne (noms des colonnes) d'un fichier CSV."""
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        reader = _csv.reader(f, delimiter=str(delimiter))
        header = next(reader, [])
    return _OktopiosList(header)


def _csv_count(path, delimiter=",", skip_header=True):
    """Compte le nombre de lignes de données dans un fichier CSV."""
    with open(str(path), newline="", encoding="utf-8-sig") as f:
        total = sum(1 for _ in f)
    return total - (1 if bool(skip_header) else 0)


# ------------------------------------------------------------------
# Helpers Table — rendu de tableaux formatés avec tabulate
# ------------------------------------------------------------------
try:
    from tabulate import tabulate as _tabulate
    _TABULATE_OK = True
except ImportError:
    _TABULATE_OK = False

_TABLE_STYLES = [
    "plain", "simple", "github", "grid", "simple_grid", "rounded_grid",
    "heavy_grid", "pipe", "orgtbl", "presto", "pretty", "psql",
    "rst", "mediawiki", "html", "tsv",
]

def _table_coerce(data):
    """Converts OktopiosList/OktopiosMap rows to plain Python lists/dicts."""
    rows = list(data) if not isinstance(data, list) else data
    result = []
    for row in rows:
        if isinstance(row, _OktopiosMap):
            result.append(dict(row))
        elif isinstance(row, _OktopiosList):
            result.append(list(row))
        elif isinstance(row, list):
            result.append(row)
        elif isinstance(row, dict):
            result.append(row)
        else:
            result.append([row])
    return result

def _table_render(data, style="simple", headers=None):
    """Returns a formatted table string from a list of maps or lists."""
    if not _TABULATE_OK:
        raise RuntimeError(
            "Le namespace Table nécessite la bibliothèque 'tabulate'.\n"
            "Installez-la avec : pip install tabulate"
        )
    rows = _table_coerce(data)
    style = str(style) if style else "simple"
    if not rows:
        return ""
    # Auto-detect headers from first row if it's a dict
    if isinstance(rows[0], dict):
        hdrs = headers if headers is not None else "keys"
    else:
        hdrs = list(headers) if headers is not None else ()
    return _tabulate(rows, headers=hdrs, tablefmt=style)

def _table_print(data, style="simple", headers=None):
    """Prints a formatted table to stdout and returns None."""
    print(_table_render(data, style, headers))
    return None

def _table_from_csv(path, style="simple", delimiter=","):
    """Reads a CSV file and returns a formatted table string."""
    rows = _csv_read(path, delimiter, has_header=True)
    return _table_render(rows, style, headers=None)

def _table_column(data, key):
    """Extracts a single column by name (str) or index (int) from data."""
    rows = _table_coerce(data)
    result = []
    for row in rows:
        if isinstance(row, dict):
            result.append(row.get(str(key), None))
        elif isinstance(row, list):
            idx = int(key)
            result.append(row[idx] if 0 <= idx < len(row) else None)
    return _OktopiosList(result)

def _table_row_count(data):
    """Returns the number of data rows."""
    rows = _table_coerce(data)
    return len(rows)

def _table_col_count(data):
    """Returns the number of columns (based on first row)."""
    rows = _table_coerce(data)
    if not rows:
        return 0
    first = rows[0]
    if isinstance(first, dict):
        return len(first)
    return len(first)

def _table_transpose(data):
    """Transposes rows and columns for list-of-lists data."""
    rows = _table_coerce(data)
    if not rows or not isinstance(rows[0], (list, _OktopiosList)):
        return _OktopiosList([])
    transposed = list(map(list, zip(*rows)))
    return _OktopiosList([_OktopiosList(r) for r in transposed])

# ------------------------------------------------------------------
# Namespace Fmt — formatage humain de nombres, durées, tailles, etc.
# Utilise uniquement la bibliothèque standard Python — aucune dépendance.
# ------------------------------------------------------------------

def _fmt_number(n, decimals=2, sep=","):
    """Formate un nombre avec séparateur de milliers et décimales."""
    try:
        n = float(n)
        decimals = int(decimals)
        sep = str(sep) if sep is not None else ","
        if decimals == 0:
            fmt = f"{int(round(n)):,}".replace(",", sep)
        else:
            fmt = f"{n:,.{decimals}f}".replace(",", "\x00").replace(".", ".").replace("\x00", sep)
        return fmt
    except Exception as e:
        return str(n)

def _fmt_percent(n, decimals=1):
    """Formate un nombre en pourcentage (0.75 → '75.0 %')."""
    try:
        n = float(n)
        decimals = int(decimals)
        return f"{n * 100:.{decimals}f} %"
    except Exception:
        return str(n)

def _fmt_currency(n, symbol="$", decimals=2):
    """Formate un nombre comme une valeur monétaire ('$ 1,234.50')."""
    try:
        n = float(n)
        decimals = int(decimals)
        symbol = str(symbol) if symbol is not None else "$"
        formatted = f"{n:,.{decimals}f}"
        return f"{symbol} {formatted}"
    except Exception:
        return str(n)

def _fmt_bytes(n):
    """Formate un nombre d'octets en notation humaine lisible (KB, MB, GB, …)."""
    try:
        n = float(n)
        for unit in ("B", "KB", "MB", "GB", "TB", "PB"):
            if abs(n) < 1024.0:
                return f"{n:.2f} {unit}"
            n /= 1024.0
        return f"{n:.2f} EB"
    except Exception:
        return str(n)

def _fmt_duration(seconds):
    """Formate une durée en secondes en texte lisible ('1h 23m 5s')."""
    try:
        seconds = int(seconds)
        if seconds < 0:
            return "0s"
        days, rem = divmod(seconds, 86400)
        hours, rem = divmod(rem, 3600)
        minutes, secs = divmod(rem, 60)
        parts = []
        if days:
            parts.append(f"{days}j")
        if hours:
            parts.append(f"{hours}h")
        if minutes:
            parts.append(f"{minutes}m")
        if secs or not parts:
            parts.append(f"{secs}s")
        return " ".join(parts)
    except Exception:
        return str(seconds)

def _fmt_plural(n, singular, plural=None):
    """Retourne la forme plurielle ou singulière selon n ('1 chat' / '3 chats')."""
    try:
        n = float(n)
        singular = str(singular)
        if plural is None:
            plural = singular + "s"
        else:
            plural = str(plural)
        word = singular if abs(n) <= 1 else plural
        if n == int(n):
            return f"{int(n)} {word}"
        return f"{n} {word}"
    except Exception:
        return str(n)

def _fmt_pad(s, width, char=" ", align="l"):
    """Aligne une chaîne dans une largeur donnée (l=gauche, r=droite, c=centre)."""
    try:
        s = str(s)
        width = int(width)
        char = str(char)[0] if char else " "
        align = str(align).lower()
        if align in ("r", "right"):
            return s.rjust(width, char)
        elif align in ("c", "center", "centre"):
            return s.center(width, char)
        else:
            return s.ljust(width, char)
    except Exception:
        return str(s)

def _fmt_truncate(s, width, suffix="…"):
    """Tronque une chaîne à width caractères en ajoutant suffix si nécessaire."""
    try:
        s = str(s)
        width = int(width)
        suffix = str(suffix) if suffix is not None else "…"
        if len(s) <= width:
            return s
        cut = max(0, width - len(suffix))
        return s[:cut] + suffix
    except Exception:
        return str(s)

def _fmt_ordinal(n):
    """Retourne la forme ordinale d'un entier en anglais (1 → '1st', 2 → '2nd', …)."""
    try:
        n = int(n)
        if 11 <= (n % 100) <= 13:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suffix}"
    except Exception:
        return str(n)

# ------------------------------------------------------------------
NativeFuncs = {
    "Math" : {
        # --- Constantes ---
        "pi": math.pi,
        "e": math.e,
        "tau": math.tau,
        "inf": math.inf,
        # --- Maths de base ---
        "abs": lambda x: abs(x),
        "round": lambda x, n=0: round(x, int(n)),
        "floor": lambda x: math.floor(x),
        "ceil": lambda x: math.ceil(x),
        "sqrt": lambda x: math.sqrt(x),
        "pow": lambda x, y: math.pow(x, y),
        "max": lambda *args: max(args),
        "min": lambda *args: min(args),
        "sum": lambda lst: sum(lst),
        "avg": lambda lst: sum(lst) / len(lst) if lst else 0,
        "sign": lambda x: (1 if x > 0 else -1 if x < 0 else 0),
        "clamp": lambda x, lo, hi: max(lo, min(hi, x)),
        # --- Maths avancées ---
        "factorial": lambda x: math.factorial(int(x)),
        "gcd": lambda a, b: math.gcd(int(a), int(b)),
        "lcm": lambda a, b: abs(int(a) * int(b)) // math.gcd(int(a), int(b)) if int(a) and int(b) else 0,
        "hypot": lambda x, y: math.hypot(x, y),
        "log": lambda x, base=None: math.log(x, base) if base is not None else math.log(x),
        "log2": lambda x: math.log2(x),
        "log10": lambda x: math.log10(x),
        "exp": lambda x: math.exp(x),
        "sin": lambda x: math.sin(x),
        "cos": lambda x: math.cos(x),
        "tan": lambda x: math.tan(x),
        "asin": lambda x: math.asin(x),
        "acos": lambda x: math.acos(x),
        "atan": lambda x: math.atan(x),
        "atan2": lambda y, x: math.atan2(y, x),
        "deg": lambda x: math.degrees(x),
        "rad": lambda x: math.radians(x),
        "isnan": lambda x: math.isnan(x),
        "isinf": lambda x: math.isinf(x),
    },
    "Time":{
        # --- Temps ---
        "sleep": lambda s: _time.sleep(float(s)),
        "time": lambda *args: _time.time(),
        "now": lambda *args: _time.time(),
        "date": lambda *args: _time.strftime("%Y-%m-%d %H:%M:%S"),
        "today": lambda *args: _time.strftime("%Y-%m-%d"),
        "ctime": lambda *args: _time.ctime(),
        "strftime": lambda fmt: _time.strftime(fmt),
        "year": lambda *args: int(_time.strftime("%Y")),
        "month": lambda *args: int(_time.strftime("%m")),
        "day": lambda *args: int(_time.strftime("%d")),
        "hour": lambda *args: int(_time.strftime("%H")),
        "minute": lambda *args: int(_time.strftime("%M")),
        "second": lambda *args: int(_time.strftime("%S")),
    },
    "String":{
        "trim": lambda s: s.strip(),
        "lstrip": lambda s: s.lstrip(),
        "rstrip": lambda s: s.rstrip(),
        "upper": lambda s: s.upper(),
        "lower": lambda s: s.lower(),
        "title": lambda s: str(s).title(),
        "length": lambda s: len(s),
        "contains": lambda s, sub: sub in s,
        "count": lambda s, sub: str(s).count(sub),
        "replace": lambda s, a, b: s.replace(a, b),
        "replaceAll": lambda s, a, b: s.replace(a, b),
        "substring": lambda s, start, end: s[int(start):int(end)],
        "toString": lambda x: str(x),
        "startswith": lambda s, prefix: s.startswith(prefix),
        "startsWith": lambda s, prefix: s.startswith(prefix),
        "endsWith": lambda s, suffix: s.endswith(suffix),
        "indexof": lambda s, sub: s.find(sub),
        "indexOf": lambda s, sub: s.find(sub),
        "lastIndexOf": lambda s, sub: s.rfind(sub),
        "isempty": lambda s: s == "",
        "isAlpha": lambda s: str(s).isalpha(),
        "isDigit": lambda s: str(s).isdigit(),
        "isAlNum": lambda s: str(s).isalnum(),
        "camelcase": lambda s: CamelCase().hump(s),
        "compareTo": lambda a, b: (a > b) - (a < b),
        "compare": lambda a, b: (a > b) - (a < b),
        "equals": lambda a, b: a == b,
        "hashCode": lambda x: hash(x),
        "capitalize": lambda s: str(s).capitalize(),
        "reverse": lambda s: str(s)[::-1],
        "repeat": lambda s, n: str(s) * int(n),
        "padStart": lambda s, width, char=" ": str(s).rjust(int(width), str(char)[0]),
        "padEnd": lambda s, width, char=" ": str(s).ljust(int(width), str(char)[0]),
        "split": lambda s, sep=" ": s.split(sep),
        "join": lambda sep, lst: str(sep).join(str(x) for x in lst),
        "format": lambda template, *args: template.format(*args),
    },
    "File":{
        # --- Fichiers & répertoires ---
        "mkdir": lambda path: os.makedirs(path, exist_ok=True),
        "rmdir": lambda path: os.rmdir(path),
        "remove": lambda path: os.remove(path),
        "rename": lambda src, dst: os.rename(src, dst),
        "exists": lambda path: os.path.exists(path),
        "isfile": lambda path: os.path.isfile(path),
        "isdir": lambda path: os.path.isdir(path),
        "read": lambda path: _file_read(path),
        "readfile": lambda path: _file_read(path),
        "write": lambda path, content: _file_write(path, content),
        "writefile": lambda path, content: _file_write(path, content),
        "append": lambda path, content: _file_append(path, content),
        "appendfile": lambda path, content: _file_append(path, content),
        "hashfile": lambda path, algo="sha256": _file_hash(path, algo),
        "zip": lambda src, dst: shutil.make_archive(dst, 'zip', src),
        "unzip": lambda zip_path, extract_to: shutil.unpack_archive(zip_path, extract_to),
        "copy": lambda src, dst: shutil.copy2(src, dst),
        "move": lambda src, dst: shutil.move(src, dst),
        "size": lambda path: os.path.getsize(path),
        "listdir": lambda path=".": os.listdir(path),
        "readlines": lambda path: _file_readlines(path),
        "basename": lambda path: os.path.basename(path),
        "dirname": lambda path: os.path.dirname(path),
        "extension": lambda path: os.path.splitext(path)[1],
        "abspath": lambda path: os.path.abspath(path),
        "join": lambda *parts: os.path.join(*parts),
    },
    "Environnement": {
        # --- Environnement ---
        "getenv": lambda var: os.getenv(var),
        "setenv": lambda var, val: os.environ.__setitem__(var, val),
        "delenv": lambda var: os.environ.__delitem__(var) if var in os.environ else None,
        "env": lambda *args: dict(os.environ),
    },
    "Processus": {
        # --- Processus ---
        "pid": lambda *args: os.getpid(),
        "ppid": lambda *args: os.getppid(),
        "run": lambda cmd: os.popen(cmd).read(),
    },
    "Identity": {
        # --- Utilisateur ---
        "user": lambda *args: os.getlogin() if hasattr(os, 'getlogin') else os.getenv("USER") or os.getenv("USERNAME"),
        "home": lambda *args: os.path.expanduser("~"),
        "is_admin": lambda *args: os.getuid() == 0 if hasattr(os, 'getuid') else ctypes.windll.shell32.IsUserAnAdmin() if os.name == 'nt' else False,
        "is_virtual_env": lambda *args: hasattr(sys, 'real_prefix') or (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix),
        "is_docker": lambda *args: os.path.exists('/.dockerenv'),
        "hostname": lambda *args: platform.node(),
        "platform": lambda *args: platform.platform(),
    },
    "System": {
        # --- Système & environnement ---
        "clear": lambda *args: os.system('cls' if os.name == 'nt' else 'clear'),
        "cls": lambda *args: os.system('cls' if os.name == 'nt' else 'clear'),
        "exit": lambda code=0: sys.exit(int(code)),
        "cwd": lambda *args: os.getcwd(),
        "cd": lambda path: os.chdir(path),
        "ls": lambda path=".": os.listdir(path),
        "sysinfo": lambda *args: {
            "os": platform.system(),
            "version": platform.version(),
            "release": platform.release(),
            "architecture": platform.architecture(),
            "machine": platform.machine(),
            "processor": platform.processor(),
            "python": platform.python_version(),
            "cwd": os.getcwd()
        },
        "whoami": lambda *args: os.getlogin() if hasattr(os, 'getlogin') else os.getenv("USER") or os.getenv("USERNAME"),
        "uid": lambda *args: os.getuid() if hasattr(os, 'getuid') else None,
        "gid": lambda *args: os.getgid() if hasattr(os, 'getgid') else None,
        "chmod": lambda path, mode: os.chmod(path, int(mode, 8)),
        "chown": lambda path, uid, gid: os.chown(path, uid, gid) if hasattr(os, 'chown') else None,
        "stat": lambda path: os.stat(path),
        "access": lambda path, mode: os.access(path, mode),  # mode = os.R_OK, os.W_OK, os.X_OK
        "uname": lambda *args: platform.uname()._asdict(),
        "cpu_count": lambda *args: os.cpu_count(),
        "loadavg": lambda *args: os.getloadavg() if hasattr(os, 'getloadavg') else None,
        "uptime": lambda *args: _time.time() - _psutil.boot_time() if _psutil else None,
        "disk_usage": lambda path=".": shutil.disk_usage(path)._asdict(),
        "memory_info": lambda *args: _psutil.virtual_memory()._asdict() if _psutil else None,
        "which": lambda cmd: shutil.which(cmd),
        "run": lambda cmd: os.popen(cmd).read(),
        "exec": lambda cmd: subprocess.run(cmd, shell=True, capture_output=True, text=True).stdout,
        #"python_version": lambda *args: platform.python_version(),
    },
    "Random":{
        "random": lambda *args: _random.random(),
        # Nombre aléatoire entier entre min et max inclus
        "randInt": lambda min_val, max_val: _random.randint(int(min_val), int(max_val)),
        # Nombre aléatoire flottant entre min et max
        "randFloat": lambda min_val, max_val: _random.uniform(float(min_val), float(max_val)),
        # Choix aleatoire dans une liste
        "choice": lambda lst: _random.choice(list(lst)),
        # Melange d'une liste (retourne une nouvelle liste melangee)
        "shuffle": lambda lst: _random.sample(list(lst), len(list(lst))),
        # Echantillon sans remise
        "sample": lambda lst, k: _random.sample(list(lst), int(k)),
        # Graine pour reproductibilite
        "seed": lambda n: _random.seed(int(n)),
        # UUID simple (hex aleatoire)
        "uuid": lambda *args: "%08x-%04x-%04x-%04x-%012x" % (
            _random.randint(0, 0xffffffff), _random.randint(0, 0xffff),
            _random.randint(0, 0xffff), _random.randint(0, 0xffff),
            _random.randint(0, 0xffffffffffff)),
    },
    "Size_STR": {
        "len": lambda x: len(x),
        "length": lambda x: len(x),
    },
    "Type": {
        # Retourne le nom de type Oktopios (pas le nom Python interne)
        "type":     lambda x: _oktype(x),
        # Prédicats de type
        "isInt":    lambda x: isinstance(x, int) and not isinstance(x, bool),
        "isFloat":  lambda x: isinstance(x, float),
        "isBool":   lambda x: isinstance(x, bool),
        "isString": lambda x: isinstance(x, str),
        "isList":   lambda x: isinstance(x, (list, _OktopiosList)),
        "isMap":    lambda x: isinstance(x, (dict, _OktopiosMap)),
        "isNull":   lambda x: x is None,
        "isNum":    lambda x: isinstance(x, (int, float)) and not isinstance(x, bool),
    },
    # ---------    # -------------------------------------------------------------------
    # List — utilitaires fonctionnels sur les listes
    # -------------------------------------------------------------------
    "List": {
        # --- Accès ---
        "head":      lambda lst: (list(lst)[0] if list(lst) else None),
        "tail":      lambda lst: _OktopiosList(list(lst)[1:]),
        "last":      lambda lst: (list(lst)[-1] if list(lst) else None),
        "init":      lambda lst: _OktopiosList(list(lst)[:-1]),
        "take":      lambda lst, n: _OktopiosList(list(lst)[:int(n)]),
        "drop":      lambda lst, n: _OktopiosList(list(lst)[int(n):]),
        "get":       lambda lst, i: list(lst)[int(i)],
        # --- Transformation ---
        "flatten":   lambda lst: _OktopiosList(_list_flatten(list(lst))),
        "unique":    lambda lst: _OktopiosList(list(dict.fromkeys(list(lst)))),
        "zip":       lambda a, b: _OktopiosList([_OktopiosList([x, y]) for x, y in zip(list(a), list(b))]),
        "unzip":     lambda lst: _OktopiosList([
                         _OktopiosList([list(p)[0] for p in list(lst)]),
                         _OktopiosList([list(p)[1] for p in list(lst)])
                     ]),
        "chunk":     lambda lst, n: _OktopiosList([
                         _OktopiosList(list(lst)[i:i + int(n)])
                         for i in range(0, len(list(lst)), int(n))
                     ]),
        "sorted":    lambda lst, rev=False: _OktopiosList(sorted(list(lst), reverse=bool(rev))),
        "reversed":  lambda lst: _OktopiosList(list(reversed(list(lst)))),
        "concat":    lambda a, b: _OktopiosList(list(a) + list(b)),
        "enumerate": lambda lst: _OktopiosList([_OktopiosList([i, v]) for i, v in enumerate(list(lst))]),
        "rotate":    lambda lst, n: _OktopiosList(_list_rotate(list(lst), int(n))),
        # --- Agrégation ---
        "sum":       lambda lst: sum(list(lst)),
        "product":   lambda lst: _list_product(list(lst)),
        "max":       lambda lst: max(list(lst)),
        "min":       lambda lst: min(list(lst)),
        "avg":       lambda lst: (sum(list(lst)) / len(list(lst))) if list(lst) else 0,
        # --- Recherche ---
        "contains":  lambda lst, x: x in list(lst),
        "indexOf":   lambda lst, x: list(lst).index(x) if x in list(lst) else -1,
        "count":     lambda lst, x: list(lst).count(x),
        # --- Ensembles ---
        "intersect": lambda a, b: _OktopiosList([x for x in list(a) if x in set(list(b))]),
        "subtract":  lambda a, b: _OktopiosList([x for x in list(a) if x not in set(list(b))]),
        "union":     lambda a, b: _OktopiosList(list(dict.fromkeys(list(a) + list(b)))),
    },
    "ValueTO": {
        "ascii":   lambda c: ord(str(c)[0]),
        "toChar":  lambda n: chr(int(n)),
        "toInt":   lambda x: int(x) if x is not None else 0,
        "toFloat": lambda x: float(x) if x is not None else 0.0,
        "toBool":  lambda x: bool(x) if x is not None else False,
    },
    "Size_T": {
        "range": lambda a, b=None: list(range(int(a), int(b))) if b is not None else list(range(int(a))),
    },
    "DataImport": {
        "readJson":   lambda path: _read_json(path),
        "readExcel":  lambda path, sheet=0: _read_excel(path, sheet),
        "readSQL":    lambda path, query="SELECT name FROM sqlite_master WHERE type='table'": _read_sql(path, query),
        "readMySQL":  lambda host, user, password, database, query, port=3306: _read_mysql(host, user, password, database, query, port),
        "writeJson":  lambda path, data: _write_json(path, data),
        "writeExcel": lambda path, data, sheet="Sheet1": _write_excel(path, data, sheet),
        "writeSQL":   lambda path, table, data: _write_sql(path, table, data),
        "writeMySQL": lambda host, user, password, database, table, data, port=3306: _write_mysql(host, user, password, database, table, data, port),
    },
    "Recognize": {
        "textNormalize":   lambda s: _text_normalize(s),
        "textSimilarity":  lambda a, b: _text_similarity(a, b),
        "facial":          lambda path: _facial_extract(path),
        "facialSimilarity":lambda fp1, fp2: _facial_similarity(fp1, fp2),
        "vocal":           lambda path, n_mfcc=20: _vocal_extract(path, n_mfcc),
        "vocalSimilarity": lambda fp1, fp2: _vocal_similarity(fp1, fp2),
    },
    "IAModule": {
        "ollama":    lambda prompt, model="llama3", host="http://localhost:11434", timeout=60, retries=2: _ia_ollama(prompt, model, host, timeout, retries),
        "deepseek":  lambda prompt, api_key=None, model=None, timeout=60, retries=2, host="http://localhost:11434": _ia_deepseek(prompt, api_key, model, timeout, retries, host),
        "starcoder": lambda prompt, api_key=None, model=None, timeout=60, retries=2, host="http://localhost:11434": _ia_starcoder(prompt, api_key, model, timeout, retries, host),
    },
    # -------------------------------------------------------------------
    # Map — utilitaires fonctionnels sur les maps (dictionnaires)
    # -------------------------------------------------------------------
    "Map": {
        # --- Introspection ---
        "keys":     lambda m: _OktopiosList(list(dict(m).keys())),
        "values":   lambda m: _OktopiosList(list(dict(m).values())),
        "entries":  lambda m: _OktopiosList([_OktopiosList([k, v]) for k, v in dict(m).items()]),
        "size":     lambda m: len(dict(m)),
        "isEmpty":  lambda m: len(dict(m)) == 0,
        "has":      lambda m, key: key in dict(m),
        # --- Accès sécurisé ---
        "get":      lambda m, key, default=None: dict(m).get(key, default),
        # --- Transformation (retourne une nouvelle map) ---
        "set":      lambda m, key, val: _OktopiosMap({**dict(m), key: val}),
        "remove":   lambda m, key: _OktopiosMap({k: v for k, v in dict(m).items() if k != key}),
        "merge":    lambda a, b: _OktopiosMap({**dict(a), **dict(b)}),
        "pick":     lambda m, keys: _OktopiosMap({k: dict(m)[k] for k in list(keys) if k in dict(m)}),
        "omit":     lambda m, keys: _OktopiosMap({k: v for k, v in dict(m).items() if k not in list(keys)}),
        # --- Conversion ---
        "fromList": lambda lst: _OktopiosMap({list(p)[0]: list(p)[1] for p in list(lst)}),
        "toList":   lambda m: _OktopiosList([_OktopiosList([k, v]) for k, v in dict(m).items()]),
        # --- Recherche ---
        "findKey":  lambda m, val: next((k for k, v in dict(m).items() if v == val), None),
        "invert":   lambda m: _OktopiosMap({v: k for k, v in dict(m).items()}),
    },
    # -------------------------------------------------------------------
    # Regex — expressions régulières
    # -------------------------------------------------------------------
    "Regex": {
        # Test si le pattern correspond quelque part dans la chaîne (booléen)
        "test":        lambda pattern, s: bool(_re.search(str(pattern), str(s))),
        # Test si la chaîne entière correspond au pattern (booléen)
        "match":       lambda pattern, s: bool(_re.fullmatch(str(pattern), str(s))),
        # Retourne la première correspondance ou null
        "search":      lambda pattern, s: (_re.search(str(pattern), str(s)).group(0)
                           if _re.search(str(pattern), str(s)) else None),
        # Retourne toutes les correspondances sous forme de liste
        "findAll":     lambda pattern, s: _re_find_all(str(pattern), str(s)),
        # Remplace les correspondances par repl
        "replace":     lambda pattern, repl, s: _re.sub(str(pattern), str(repl), str(s)),
        # Remplace exactement n occurrences (0 = toutes)
        "replaceN":    lambda pattern, repl, s, n=0: _re.sub(str(pattern), str(repl), str(s), count=int(n)),
        # Découpe la chaîne selon le pattern
        "split":       lambda pattern, s: _OktopiosList(_re.split(str(pattern), str(s))),
        # Groupes de capture de la première correspondance (liste ordonnée)
        "groups":      lambda pattern, s: _re_groups(str(pattern), str(s)),
        # Groupes nommés de la première correspondance (map)
        "namedGroups": lambda pattern, s: _re_named_groups(str(pattern), str(s)),
        # Compte le nombre de correspondances
        "count":       lambda pattern, s: len(_re.findall(str(pattern), str(s))),
        # Échappe les caractères spéciaux d'une chaîne pour usage dans un pattern
        "escape":      lambda s: _re.escape(str(s)),
    },
    # -------------------------------------------------------------------
    # Date — manipulation et arithmétique de dates
    # Toutes les dates entrantes/sortantes sont des chaînes ISO "YYYY-MM-DD"
    # (ou formats courants DD/MM/YYYY, MM/DD/YYYY, etc.).
    # -------------------------------------------------------------------
    "Date": {
        # Parsing / formatage
        "parse":         lambda s, fmt="%Y-%m-%d": _date_parse(s, fmt),
        "format":        lambda d, fmt="%d/%m/%Y": _date_format(d, fmt),
        # Arithmétique
        "add":           lambda d, n, unit="days": _date_add(d, n, unit),
        "diff":          lambda d1, d2, unit="days": _date_diff(d1, d2, unit),
        # Comparaison  (-1 = avant, 0 = égal, 1 = après)
        "compare":       lambda d1, d2: _date_compare(d1, d2),
        "isBefore":      lambda d1, d2: _date_compare(d1, d2) < 0,
        "isAfter":       lambda d1, d2: _date_compare(d1, d2) > 0,
        "isEqual":       lambda d1, d2: _date_compare(d1, d2) == 0,
        # Informations calendaires
        "weekday":       lambda d: _date_weekday(d),
        "weekdayName":   lambda d, lang="en": _date_weekday_name(d, lang),
        "isLeapYear":    lambda year: _date_is_leap_year(int(year)),
        "daysInMonth":   lambda year, month: _date_days_in_month(int(year), int(month)),
        # Conversion timestamp Unix
        "toTimestamp":   lambda d: _date_to_timestamp(d),
        "fromTimestamp": lambda ts: _date_from_timestamp(ts),
    },
    # -------------------------------------------------------------------
    # Json — manipulation de JSON en mémoire
    # parse/stringify, accès par chemin pointé, merge profond
    # -------------------------------------------------------------------
    "Json": {
        # Désérialise une chaîne JSON en valeur Oktopios (map/liste/primitif)
        "parse":        lambda s: _json_parse(str(s)),
        # Sérialise une valeur Oktopios en chaîne JSON compacte
        "stringify":    lambda v, indent=None: _json_stringify(v, indent),
        # Sérialise avec indentation (pretty-print)
        "pretty":       lambda v: _json_stringify(v, 2),
        # Lit la valeur à un chemin "a.b.c" dans un objet imbriqué
        "get":          lambda obj, path, default=None: _json_path_get(obj, str(path), default),
        # Retourne un nouvel objet avec la valeur modifiée au chemin donné
        "set":          lambda obj, path, val: _json_path_set(obj, str(path), val),
        # Vérifie si le chemin existe dans l'objet
        "has":          lambda obj, path: _json_path_has(obj, str(path)),
        # Merge profond de deux objets JSON (b écrase a en cas de conflit)
        "merge":        lambda a, b: _json_deep_merge(a, b),
        # Charge un fichier JSON et le retourne comme valeur Oktopios
        "fromFile":     lambda path: _json_from_file(str(path)),
        # Écrit une valeur Oktopios dans un fichier JSON (pretty par défaut)
        "toFile":       lambda path, v, indent=2: _json_to_file(str(path), v, int(indent)),
        # Valide qu'une chaîne est du JSON valide (booléen)
        "isValid":      lambda s: _json_is_valid(str(s)),
        # Retourne les clés de premier niveau d'un objet JSON
        "keys":         lambda obj: _OktopiosList(list(dict(obj).keys())) if isinstance(obj, (dict, _OktopiosMap)) else _OktopiosList([]),
    },

    # -------------------------------------------------------------------
    # Http — requêtes HTTP GET / POST / PUT / PATCH / DELETE
    # Chaque méthode retourne un map : { status, ok, body, headers }
    # Nécessite : pip install requests   (ou pip install oktopios[ia])
    # -------------------------------------------------------------------
    "Http": {
        # Méthodes HTTP principales
        "get":    lambda url, headers=None, timeout=30: _http_get(str(url), headers, timeout),
        "post":   lambda url, body=None, headers=None, json=None, timeout=30: _http_post(str(url), body, headers, json, timeout),
        "put":    lambda url, body=None, headers=None, json=None, timeout=30: _http_put(str(url), body, headers, json, timeout),
        "patch":  lambda url, body=None, headers=None, json=None, timeout=30: _http_patch(str(url), body, headers, json, timeout),
        "delete": lambda url, headers=None, timeout=30: _http_delete(str(url), headers, timeout),
        # Accesseurs sur la réponse retournée
        "status":  lambda response: _http_status(response),
        "ok":      lambda response: _http_ok(response),
        "body":    lambda response: _http_body(response),
        "json":    lambda response: _http_json(response),
        "headers": lambda response: _http_headers(response),
    },

    # ------------------------------------------------------------------
    # Namespace Set — ensemble mathématique (valeurs uniques, non ordonnées)
    # Représentation interne : OktopiosMap { élément -> True }
    # Toutes les fonctions sont pures (immutables) — elles retournent
    # un nouveau Set sans modifier l'original.
    # ------------------------------------------------------------------
    "Set": {
        # Création
        "create":     lambda lst=None: _OktopiosMap({k: True for k in (lst if lst is not None else [])}),
        "fromList":   lambda lst: _OktopiosMap({k: True for k in (lst if lst is not None else [])}),
        # Lecture
        "has":        lambda s, x: x in s,
        "size":       lambda s: len(s),
        "isEmpty":    lambda s: len(s) == 0,
        "toList":     lambda s: _OktopiosList(list(s.keys())),
        # Modification (retourne un nouveau Set)
        "add":        lambda s, x: _OktopiosMap({**s, x: True}),
        "remove":     lambda s, x: _OktopiosMap({k: True for k in s if k != x}),
        "clear":      lambda s: _OktopiosMap({}),
        # Opérations ensemblistes
        "union":      lambda a, b: _OktopiosMap({**{k: True for k in a}, **{k: True for k in b}}),
        "intersect":  lambda a, b: _OktopiosMap({k: True for k in a if k in b}),
        "diff":       lambda a, b: _OktopiosMap({k: True for k in a if k not in b}),
        "symDiff":    lambda a, b: _OktopiosMap({
            **{k: True for k in a if k not in b},
            **{k: True for k in b if k not in a}
        }),
        # Prédicats relationnels
        "isSubset":   lambda a, b: all(k in b for k in a),
        "isSuperset": lambda a, b: all(k in a for k in b),
        "isDisjoint": lambda a, b: not any(k in b for k in a),
        "equals":     lambda a, b: set(a.keys()) == set(b.keys()),
    },

    # ------------------------------------------------------------------
    # Namespace Queue — file d'attente FIFO (premier entré, premier sorti)
    # Représentation interne : OktopiosMap { "_t": "Q", "_d": [items] }
    # Les mutations (enqueue/dequeue/clear) agissent directement sur la
    # liste interne — la Queue est donc un objet mutable.
    # ------------------------------------------------------------------
    "Queue": {
        # Création
        "create":   _queue_create,
        "fromList": lambda lst: _queue_create(lst),
        # Mutation (FIFO)
        "enqueue":  _queue_enqueue,   # ajoute un élément en fin de file
        "dequeue":  _queue_dequeue,   # retire et retourne l'élément de tête
        # Lecture sans modification
        "peek":     _queue_peek,      # retourne l'élément de tête sans le retirer
        "size":     _queue_size,      # nombre d'éléments dans la file
        "isEmpty":  _queue_isEmpty,   # true si la file est vide
        "contains": _queue_contains,  # true si val est dans la file
        # Conversion / réinitialisation
        "toList":   _queue_toList,    # retourne une OktopiosList (tête en premier)
        "clear":    _queue_clear,     # vide la file
    },

    # ------------------------------------------------------------------
    # Namespace Stack — pile LIFO (dernier entré, premier sorti)
    # Représentation interne : OktopiosMap { "_t": "S", "_d": [items] }
    # Le sommet (top) est le dernier élément de la liste interne.
    # Les mutations (push/pop/clear) agissent directement sur la liste.
    # ------------------------------------------------------------------
    "Stack": {
        # Création
        "create":   _stack_create,
        "fromList": lambda lst: _stack_create(lst),
        # Mutation (LIFO)
        "push":     _stack_push,      # empile un élément au sommet
        "pop":      _stack_pop,       # dépile et retourne l'élément du sommet
        # Lecture sans modification
        "peek":     _stack_peek,      # retourne l'élément du sommet sans le retirer
        "size":     _stack_size,      # nombre d'éléments dans la pile
        "isEmpty":  _stack_isEmpty,   # true si la pile est vide
        "contains": _stack_contains,  # true si val est dans la pile
        # Conversion / réinitialisation
        "toList":   _stack_toList,    # retourne une OktopiosList (sommet en premier)
        "clear":    _stack_clear,     # vide la pile
    },

    # ------------------------------------------------------------------
    # Namespace Hash — hachage cryptographique & encodage Base64
    # Toutes les fonctions utilisent la bibliothèque standard Python :
    # hashlib, hmac, base64 — aucune dépendance supplémentaire requise.
    # ------------------------------------------------------------------
    "Hash": {
        # Algorithmes de hachage — retournent un digest hexadécimal (str)
        "md5":      lambda s: _hash_digest("md5",    s),
        "sha1":     lambda s: _hash_digest("sha1",   s),
        "sha256":   lambda s: _hash_digest("sha256", s),
        "sha512":   lambda s: _hash_digest("sha512", s),
        # HMAC — code d'authentification de message
        # hmac(key, msg, algo?)  — algo par défaut : "sha256"
        "hmac":     lambda key, msg, algo="sha256": _hash_hmac(key, msg, algo),
        # Encodage / décodage Base64 standard (RFC 4648 §4)
        "b64encode":    lambda s: _hash_b64encode(s),
        "b64decode":    lambda s: _hash_b64decode(s),
        # Encodage / décodage Base64 URL-safe (RFC 4648 §5 — pour JWT, URLs)
        "b64urlEncode": lambda s: _hash_b64url_encode(s),
        "b64urlDecode": lambda s: _hash_b64url_decode(s),
        # Comparaison en temps constant (résistant aux attaques temporelles)
        "compare":  lambda h1, h2: _hash_compare(h1, h2),
    },

    # ------------------------------------------------------------------
    # Namespace Stats — statistiques descriptives (stdlib uniquement)
    # Toutes les fonctions utilisent uniquement la bibliothèque standard
    # Python (statistics, math) — aucune dépendance externe requise.
    # ------------------------------------------------------------------
    "Stats": {
        # Mesures de tendance centrale
        "mean":        lambda lst: _stats_mean(lst),
        "median":      lambda lst: _stats_median(lst),
        "modeOf":      lambda lst: _stats_mode(lst),
        "geomean":     lambda lst: _stats_geomean(lst),
        "harmean":     lambda lst: _stats_harmean(lst),
        # Mesures de dispersion
        "variance":    lambda lst, pop=False: _stats_variance(lst, pop),
        "stddev":      lambda lst, pop=False: _stats_stddev(lst, pop),
        "range":       lambda lst: _stats_range(lst),
        "iqr":         lambda lst: _stats_iqr(lst),
        "mad":         lambda lst: _stats_mad(lst),
        # Quantiles et centiles
        "quartiles":   lambda lst: _stats_quartiles(lst),
        "percentile":  lambda lst, p: _stats_percentile(lst, p),
        # Normalisation et scores
        "normalize":   lambda lst: _stats_normalize(lst),
        "zscore":      lambda lst: _stats_zscore(lst),
        # Relations entre deux séries
        "covariance":  lambda a, b: _stats_covariance(a, b),
        "correlation": lambda a, b: _stats_correlation(a, b),
        # Utilitaires
        "sum":         lambda lst: float(sum(lst)),
        "min":         lambda lst: float(min(lst)),
        "max":         lambda lst: float(max(lst)),
        "size":        lambda lst: len(lst),
        "describe":    lambda lst: _stats_describe(lst),
    },

    # ------------------------------------------------------------------
    # Namespace Path — manipulation de chemins de fichiers (stdlib uniquement)
    # Toutes les fonctions utilisent os.path et pathlib de la bibliothèque
    # standard Python — aucune dépendance supplémentaire requise.
    # ------------------------------------------------------------------
    "Path": {
        # Jointure de composantes de chemin (cross-platform)
        "join":      lambda *parts: os.path.join(*[str(p) for p in parts]),
        # Répertoire parent d'un chemin ("a/b/c.txt" → "a/b")
        "dirname":   lambda p: os.path.dirname(str(p)),
        # Nom de fichier avec extension ("a/b/c.txt" → "c.txt")
        "basename":  lambda p: os.path.basename(str(p)),
        # Nom de fichier SANS extension ("a/b/c.txt" → "c")
        "stem":      lambda p: os.path.splitext(os.path.basename(str(p)))[0],
        # Extension du fichier, avec le point ("a/b/c.txt" → ".txt")
        "ext":       lambda p: os.path.splitext(str(p))[1],
        # Chemin absolu résolu depuis le répertoire courant
        "abs":       lambda p: os.path.abspath(str(p)),
        # Chemin normalisé (résout "..", "." et les séparateurs doubles)
        "normalize": lambda p: os.path.normpath(str(p)),
        # Vrai si le chemin est absolu
        "isAbs":     lambda p: os.path.isabs(str(p)),
        # Séparation en [répertoire, fichier] → liste à deux éléments
        "split":     lambda p: _OktopiosList(list(os.path.split(str(p)))),
        # Séparation en [racine, extension] → liste à deux éléments
        "splitExt":  lambda p: _OktopiosList(list(os.path.splitext(str(p)))),
        # Répertoire de travail courant
        "cwd":       lambda *a: os.getcwd(),
        # Répertoire personnel de l'utilisateur (~)
        "home":      lambda *a: os.path.expanduser("~"),
        # Développe "~" et les variables d'environnement dans un chemin
        "expand":    lambda p: os.path.expandvars(os.path.expanduser(str(p))),
        # Vérifie si le chemin existe (fichier ou dossier)
        "exists":    lambda p: os.path.exists(str(p)),
        # Vrai si c'est un fichier régulier
        "isFile":    lambda p: os.path.isfile(str(p)),
        # Vrai si c'est un répertoire
        "isDir":     lambda p: os.path.isdir(str(p)),
        # Taille du fichier en octets
        "size":      lambda p: os.path.getsize(str(p)) if os.path.exists(str(p)) else -1,
        # Liste les entrées d'un répertoire (noms seulement, pas les sous-arbres)
        "listdir":   lambda p=".": _OktopiosList(os.listdir(str(p))),
        # Chemin relatif de p par rapport à start (défaut : répertoire courant)
        "relpath":   lambda p, start=".": os.path.relpath(str(p), str(start)),
    },

    # ------------------------------------------------------------------
    # Namespace Csv — lecture, écriture et conversion CSV (stdlib uniquement)
    # Toutes les fonctions utilisent uniquement le module `csv` de la
    # bibliothèque standard Python — aucune dépendance externe requise.
    # ------------------------------------------------------------------
    "Csv": {
        # Lecture d'un fichier CSV → liste de maps (has_header=true) ou liste de listes
        "read":       lambda path, delimiter=",", has_header=True: _csv_read(path, delimiter, has_header),
        # Écriture de données dans un fichier CSV (liste de maps ou liste de listes)
        "write":      lambda path, data, delimiter=",", header=None: _csv_write(path, data, delimiter, header),
        # Analyse d'une chaîne CSV → liste de maps ou liste de listes
        "parse":      lambda text, delimiter=",", has_header=True: _csv_parse(text, delimiter, has_header),
        # Conversion de données en chaîne CSV
        "stringify":  lambda data, delimiter=",", header=None: _csv_stringify(data, delimiter, header),
        # Premières n lignes d'un fichier CSV (par défaut 5) → liste de listes
        "head":       lambda path, n=5, delimiter=",": _csv_head(path, n, delimiter),
        # Noms des colonnes (première ligne) → liste de chaînes
        "columns":    lambda path, delimiter=",": _csv_columns(path, delimiter),
        # Nombre de lignes de données (hors en-tête par défaut)
        "count":      lambda path, delimiter=",", skip_header=True: _csv_count(path, delimiter, skip_header),
    },

    # ------------------------------------------------------------------
    # Namespace Table — rendu de tableaux formatés en texte
    # Utilise la bibliothèque `tabulate` (déjà incluse dans les dépendances).
    # Supporte de nombreux styles : plain, simple, grid, pipe, github, rst …
    # ------------------------------------------------------------------
    "Table": {
        # Formate data (liste de maps ou liste de listes) en chaîne tabulaire
        "render":    lambda data, style="simple", headers=None: _table_render(data, style, headers),
        # Affiche directement la table dans le terminal
        "print":     lambda data, style="simple", headers=None: _table_print(data, style, headers),
        # Liste des styles de rendu disponibles → liste de chaînes
        "styles":    lambda *a: _OktopiosList(_TABLE_STYLES),
        # Lit un fichier CSV et retourne une chaîne de table formatée
        "fromCsv":   lambda path, style="simple", delimiter=",": _table_from_csv(path, style, delimiter),
        # Extrait une colonne par son nom ou son index → liste de valeurs
        "column":    lambda data, key: _table_column(data, key),
        # Retourne le nombre de lignes de données
        "rowCount":  lambda data: _table_row_count(data),
        # Retourne le nombre de colonnes
        "colCount":  lambda data: _table_col_count(data),
        # Transpose lignes ↔ colonnes (liste de listes uniquement)
        "transpose": lambda data: _table_transpose(data),
    },

    # ------------------------------------------------------------------
    # Namespace Fmt — formatage humain de nombres, durées, tailles, etc.
    # Utilise uniquement la bibliothèque standard Python — aucune dépendance.
    # ------------------------------------------------------------------
    "Fmt": {
        # Formate un nombre avec séparateur de milliers (ex. 1234567.8 → "1,234,567.80")
        "number":   lambda n, decimals=2, sep=",": _fmt_number(n, decimals, sep),
        # Formate en pourcentage (ex. 0.753 → "75.3 %")
        "percent":  lambda n, decimals=1: _fmt_percent(n, decimals),
        # Formate en monnaie (ex. 9.5 → "$ 9.50")
        "currency": lambda n, symbol="$", decimals=2: _fmt_currency(n, symbol, decimals),
        # Taille en octets lisible (ex. 1536000 → "1.46 MB")
        "bytes":    lambda n: _fmt_bytes(n),
        # Durée en secondes lisible (ex. 3665 → "1h 1m 5s")
        "duration": lambda s: _fmt_duration(s),
        # Pluralisation (ex. Fmt.plural(3, "chat") → "3 chats")
        "plural":   lambda n, singular, plural=None: _fmt_plural(n, singular, plural),
        # Alignement dans une largeur (l=gauche, r=droite, c=centré)
        "pad":      lambda s, width, char=" ", align="l": _fmt_pad(s, width, char, align),
        # Troncature avec suffixe (ex. "Bonjour le monde" → "Bonj…")
        "truncate": lambda s, width, suffix="…": _fmt_truncate(s, width, suffix),
        # Ordinal anglais (ex. 3 → "3rd", 11 → "11th")
        "ordinal":  lambda n: _fmt_ordinal(n),
    },

}
